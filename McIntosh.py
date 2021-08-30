# -*- coding: utf-8 -*-
"""

  
This is a code for the CLSE 202 Project 3 written by Tara McIntosh and Sravya Dhavala 

This program calculates density at different pressure- temperature isotherms using the
peng robinson equation of state as well as critical data from 
"Chemical, Biomedical and Engineering Thermodynamics" by Sandler
"""
import openpyxl  #I use openpyxl throughout to work in excel
def save(k): #this is a subroutine I call later on in  my main to save it 
    #9f from the project list of what we have to do 
   project3 = openpyxl.load_workbook("Project3_Group8 - Tara McIntosh - Sravya Dhavala.xlsx" ) #tells it to load the workbook
   k = project3.save("Project3_Group8 - Tara McIntosh - Sravya Dhavala.xlsx") #defines a variable to save it 
   return k

import numpy as np #this is used for the FINDZ cubic solver, useful for arrays in python
def FINDZ(RA1,RA2,RA3,RA4,NNN):
    A1 = ((3 * RA3)-(RA2**2))/3
    B1 = ((2*RA2**3)-(9*RA2 * RA3) + (27*RA4))/27
    test1 = abs(A1**3)/27
    test2 = (B1**2)/4
    Pi = 3.141559265
    if A1 < 0 and test1 > test2:
        CO = 2 * (((-A1)/3)**0.5)
        theta = np.arccos((3*B1)/(A1*CO))/3
        root = [1,1,1]
        root[0] = CO * np.cos(theta) - (RA2 / 3)
        root[1] = CO * np.cos(theta + (2 * Pi) / 3) - (RA2 / 3)
        root[2] = CO * np.cos(theta + (4 * Pi) / 3) - (RA2 / 3)
        for i in range (0,2):
            if root[i] < 0 and NNN == 1:
                root[i] = 1000000000
            #Max root is the vapor and Min root is the liquid.\n",
            if NNN == 0:
                Z = max(root[0], root[1], root[2])
            if NNN == 1:
                Z = min(root[0], root[1], root[2])
                
    else:
        DD = (test2 + (A1 ** 3) / 27) ** (0.5)
        AL = 1
        ALL = 1
        test3 = (-B1) / 2 + DD
        if test3 < 0: 
              AL = -1
        test3 = abs(test3)
        A2 = AL * (test3 ** (1 / 3))
        test4 = ((-B1) / 2) - DD
        if test4 < 0:
              ALL = -1
        test4 = abs(((-B1) / 2) - DD)
        B2 = ALL * (test4 ** (1 / 3))
        Z = A2 + B2 - (RA2 / 3)
        if test4 < 0.0001: 
              return(Z)   
        test5 = abs(1 - abs(A2 / B2))
        if test5 < 0.0005:
              Z = (-1 * ((A2 + B2) / 2)) - (RA2 / 3)
    return(Z)
print(FINDZ)
def main():
    
    project3 = openpyxl.load_workbook("Project3_Group8 - Tara McIntosh - Sravya Dhavala.xlsx" ) #loads the wb in the main routine
    sheet = project3['PREOS Data'] #selects the first sheet in the workbook
    print("Note: Starting pressure must be between 0 and 100bar for this project, and starting temperature selected must be between -10 and 10 degrees Celsius")
    print("Note: Enter either Methane, ethane, or Propane as compound for this program")
    import math
    #^^^ these two notes remind the user of the conditions this program works under, only for propane and butane 
    #it also works while pressure is between 0 and 100 bar 
    
    #C1 = str(input("Enter compound name:")) #inputs compound data
    #if C1 in ['Methane','methane']:
    # C2 = "Methane"
    #  C3 = "Methane Density (g/cc)" #used later on to add a header depending on # of isotherms
    Pcm =  46.1 #Critical data from Nist (bar)
    Tcm = 190.6 #kelvin
    wm = 0.008
    Mwm = 16.043
    R = 83.14
       # a1 = 25.460
       # b1 = (1.519*10**(-2))
       # c1 = (-0.715*10**(-5))
       # d1 = (1.311*10**(-9))
    #else: 
       # if C1 in ['Ethane','ethane']: #inputs compound data
           # C2 = "Ethane"
           # C3 = "Ethane Density (g/cc)"
    Pce = 48.84
    Tce = 305.4
    we = 0.098
    Mwe = 30.070
    R = 83.14
       # else: 
         #   if C1 in ['Propane','propane']: #inputs compound data
             #   C2 = "Propane"
             #   C3 = "Propane Density (g/cc)"
    Pcp = 42.46 
    Tcp = 369.8
    wp = 0.152
    Mwp = 44.097
    R = 83.14
          #  else: 
           #     print("Only Methane, propane, or ethane may be selected in this program!!") 
                #This stops the user from entering a compound like ethylene or H2O
                #as we don't have data for that
  
    n = float(input("Enter the number of isotherms:")) #allows user to select any # of isotherms they want
    n1 = int(n) #turns n into an integer so it may be used in loop calculations
    n2 = (3*n1+1) #this is used later on to loop through cells/columns/rows in excel
    
    
    P = float(input("Enter the starting pressure (bar):")) 
    if P <= 0 or P >= 32:
        print ("Error! Starting pressures must be between 0 and 10bar") #limits the user to select only a pressure in this range
        
    T = float(input("Enter the starting temperature (C):")) 
    T1 = (T+273.15)
    Tf = (T+n)
    Tc1 = (Tce)
    if T <= 0 or T>=101:
        print ("Error!, temperatures must be between -10 and 10C") #again adds a limit
    else: #this next section only inputs the 'looks' of the spreadsheet, this is NOT included in calculations, it is just so that
        #when the user is looking at the spreadsheet they can see at what conditions the calculations are occuring 
        Tloop1= T
        Tloop2= 0    
        while Tloop1 <=(n-1)+T or Tloop2<=(n-1):
            sheet.cell(row=16,column=1+Tloop2).value= Tloop1
            sheet.cell(row=15,column=1+Tloop2).value= "T" + " (Celsius)" 
            Tloop1+=1
            Tloop2+=3
        for j in range(2,n2,3): #tells which column to put pressure values in
            Ploop = P
            e = 0
            while Ploop <=20:
                sheet.cell(row=16+e,column=j).value = Ploop
                sheet.cell(row=15,column=j).value = "Pressure (bar)"
                Ploop+=10
                e+=1 
     #this next section is the actual PREOS density calculation
    Tloop = T1
    for Tloop in range(int(T1),int(T1+n),1):
        k1 = ((Tloop-int(T1))/1 +1) *3  #a k counter dependent upon TLOOP
        e = 0
        Ploop = P
        while Ploop<=20:
            NNN = 0  #for vapor species NNN = 0 
            Trm = (Tloop)/Tcm
            Tre = (Tloop)/Tce
            Trp = (Tloop)/Tcp
            
            KBIME = -0.003 #binary parameter for methane and ethane from textbook
            KBIMP = 0.016 #binary parameter for methane and propane from textbook
            KBIPE = 0.001 #binary parameter for ethane and propane from textbook
            
            ym = 0.7
            ye = 0.2 
            yp = 0.1
            
            bm = (0.07780*((R*Tcm)/(Pcm))) #this is the b in PREOS
            be = (0.07780*((R*Tce)/(Pce)))#correct
            bp = (0.07780*((R*Tcp)/(Pcp)))#correct
            
            bmix = (bm*ym) + (be*ye) + (yp*bp) #CHECK THIS WITH HIM 
            
            alpham = (1+(0.37464+(1.54226*wm)-(0.26992*(wm**2)))*(1-Trm**0.5))**2
            alphae = (1+(0.37464+(1.54226*we)-(0.26992*(we**2)))*(1-Tre**0.5))**2
            alphap = (1+(0.37464+(1.54226*wp)-(0.26992*(wp**2)))*(1-Trp**0.5))**2
            am = ((0.45724)*(((R**2)*(Tcm**2))/Pcm))*(alpham) #this is the a in PREOS THIS IS CORRECT
            ap = ((0.45724)*(((R**2)*(Tcp**2))/Pcp))*(alphap) #this is the a in PREOS THIS IS CORRECT
            ae = ((0.45724)*(((R**2)*(Tce**2))/Pce))*(alphae) #this is the a in PREOS
            
            ame = ((am*ae)**(0.5))*(1-(KBIME))
            amm = ((am*am)**(0.5))*(1)
            app = ((ap*ap)**(0.5))*(1)
            aee = ((ae*ae)**(0.5))*(1)
            amp = ((am*ap)**(0.5))*(1-(KBIMP))
            ape = ((ap*ae)**(0.5))*(1-(KBIPE))
            
            amix = 2*(ame*ym*ye)+2*((amp*yp*ym))+2*(ape*yp*ye)+(amm*ym*ym)+(app*yp*yp)+(aee*ye*ye) #is there meant to be a 2 in here ?
            
            #^^^^^CHECK WITH HIM 
            
            #k = 0.37464 + 1.542268*w - 0.26992*(w**(2))
            AA = (amix*Ploop)/((R**2)*(Tloop**2))
            BB = ((bmix*Ploop)/(R*Tloop))
            RA1 = 1
            RA2 = BB-1
            RA3 = (AA-2*BB - 3*BB**2)
            RA4 =( BB**3 +BB**2 - AA*BB)
            #Zmix = round(FINDZ(RA1,RA2,RA3,RA4,NNN),5) #calls FINDZ subroutine
            Zmix = FINDZ(RA1,RA2,RA3,RA4,NNN) #calls FINDZ subroutine
            FugacityM = ym*Ploop*math.exp(((((bm)/bmix)*(Zmix-1)) - math.log(Zmix-((bmix*Ploop)/(R*Tloop))) - ((amix)/((2*((2)**0.5))*bmix*R*Tloop)*(((2*((2*ye*ame)+(2*yp*amp)))/(amix))-((bm)/(bmix)))*math.log((Zmix+(1+(2**0.5)*((bmix*Ploop)/(R*Tloop))))/(Zmix+(1-(2**0.5)*((bmix*Ploop)/(R*Tloop))))))))
            FugacityE = ye*Ploop*math.exp(((((be)/bmix)*(Zmix-1)) - math.log(Zmix-((bmix*Ploop)/(R*Tloop))) - ((amix)/((2*((2)**0.5))*bmix*R*Tloop)*(((2*((2*ym*ame)+(2*yp*ape)))/(amix))-((be)/(bmix)))*math.log((Zmix+(1+(2**0.5)*((bmix*Ploop)/(R*Tloop))))/(Zmix+(1-(2**0.5)*((bmix*Ploop)/(R*Tloop))))))))
            FugacityP = yp*Ploop*math.exp(((((bp)/bmix)*(Zmix-1)) - math.log(Zmix-((bmix*Ploop)/(R*Tloop))) - ((amix)/((2*((2)**0.5))*bmix*R*Tloop)*(((2*((2*ym*amp)+(2*ye*ape)))/(amix))-((bp)/(bmix)))*math.log((Zmix+(1+(2**0.5)*((bmix*Ploop)/(R*Tloop))))/(Zmix+(1-(2**0.5)*((bmix*Ploop)/(R*Tloop))))))))
            #CHECK SUMMATION IN FUGACITY WITH HIM 
            #switch out program for another compound to check your values 
            
            #FINDZ1 = [RA1,RA2,RA3,RA4];
            #Z=np.roots(FINDZ1)
            #Z1 = Z[np.isreal(Z)]
            #Z3 = Z1[np.isreal(Z1)].real[2]
            #print("The following information is for oxygen in the liquid species!")
            print("Z is:", Zmix)
            print("Fugacity of Vapor Methane is", FugacityM)
            print("Fugacity of Vapor Ethane is", FugacityE)
            print("Fugacity of Vapor Propane is", FugacityP)
            print("Amix is:", amix)
            print("Bmix is:", bmix)
            #Density = ((Ploop)*Mw)/(R*(Tloop)*Z)
            #Density1 = float(round(Density,3)) #rounds the density to 3 decimal places
            
            #sheet.cell(row=16+e,column=k1).value = Density1 
            #alpha1 = (1+k*(1-((Tloop/Tc)**0.5)))
            "pure component check"
            AAm = (am*Ploop)/((R**2)*(Tloop**2))
            BBm = ((bm*Ploop)/(R*Tloop))
            RA1 = 1
            RA2m = BBm-1
            RA3m = (AAm-2*BBm - 3*BBm**2)
            RA4m =( BBm**3 +BBm**2 - AAm*BBm)
            Zm = FINDZ(RA1,RA2m,RA3m,RA4m,NNN)
            fugacitym = Ploop*math.exp((Zm-1)-math.log(Zm-BBm)-((AAm)/(2*(2**0.5)*BBm))*math.log((Zm+(1+(2**0.5))*BBm)/(Zm+(1-(2**0.5))*BBm)))
            fm = fugacitym*ym
            print("Methane check", fm)
            
            AAp = (ap*Ploop)/((R**2)*(Tloop**2))
            BBp = ((bp*Ploop)/(R*Tloop))
            RA1 = 1
            RA2p = BBp-1
            RA3p = (AAp-2*BBp - 3*BBp**2)
            RA4p =( BBp**3 +BBp**2 - AAp*BBp)
            Zp = FINDZ(RA1,RA2p,RA3p,RA4p,NNN)
            fugacityp = Ploop*math.exp((Zp-1)-math.log(Zp-BBp)-((AAp)/(2*(2**0.5)*BBp))*math.log((Zp+(1+(2**0.5))*BBp)/(Zp+(1-(2**0.5))*BBp)))
            fp = fugacityp*yp
            
            
            AAe = (ae*Ploop)/((R**2)*(Tloop**2))
            BBe = ((be*Ploop)/(R*Tloop))
            
            RA1 = 1
            RA2e = BBe-1
            RA3e = (AAe-2*BBe - 3*BBe**2)
            RA4e =( BBe**3 +BBe**2 - AAe*BBe)
            Ze = FINDZ(RA1,RA2e,RA3e,RA4e,NNN)
            Be = (Ploop*be)/(R*Tloop)
            
            fugacitye = Ploop*math.exp((Ze-1)-math.log(Ze-Be)-((AAe)/(2*(2**0.5)*Be))*math.log((Ze+(1+(2**0.5))*Be)/(Ze+(1-(2**0.5))*Be)))
            fe = fugacitye*ye
            print("Ethane check", fe) #equals partial molar fugacity
            
            print("Propane check", fp)
            
            #tells the program where to put the density values
           # V = round(((Z*(83.14/1000000)*Tloop)/Ploop),9)
           # print("The molar volume in m3/mol is:", V)
           # B = (Ploop*b)/(R*Tloop)
            #H2departure = ((8.314)*Tc*((Tr*(Z-1))-((2.078*(1+k))*(alpha**0.5)*math.log((Z+(2.414*B))/(Z-(0.414*B))))))
           
            #H2CP = (a1*((Tloop)-(298.15))) + (b1*((Tloop**2)-(298.15**2)))/(2) + (c1*((Tloop**3)-(298.15**3)))/(3) + (d1*((Tloop**4)-(298.15**4)))/(4)
            #print("The molar enthalpy in J/mol is:", H2departure + H2CP)
            
        
            #this calculates the volume that the component takes up
            #S2departure = (8.314)*(math.log(Z-B)-(2.078*k*(((1+k)/(Tr**0.5))-k))*math.log((Z+(2.414*B))/(Z-(0.414*B))))
            #S2CP = a1*(math.log(Tloop/298.15)) + (b1*((Tloop)-(298.15))) + (c1*((Tloop**2)-(298.15**2)))/(2) + (d1*((Tloop**3)-(298.15**3)))/(3) 
            #S23 = 8.314*(math.log(Ploop/1)) 
            #print("The molar entropy in J/(mol*K) is:", S2departure + S2CP - S23)
            
            #print("The Fugacity of the Species is (bar):", fugacity)
            #Gibbs = (H2departure+H2CP)-(Tloop)*(S2departure+S2CP-S23)
            #print("The Gibbs Free Energy in J/mol is:", Gibbs)
            
            #this V equation is for PV=ZRT, this equation assumes the species is acting ideally, which it isn't so there is error there
            #this is only applied to the tank with propane
            #a higher temperature and a higher pressure yield a higher volume of the tank being taken up
            #sheet.cell(row=15,column=k1).value = C3
            #sheet.cell(row=27+e,column=k1).value = V
            sheet.cell(row=26,column=k1).value = "Volume of Compound (m3/mol)"
            sheet.cell(row=26,column=k1+1).value = "Enthalpy of Compound"
            #'if V>800: #this says if the volume exceeds 800m^3 it cant be stored in the tank'
            #'sheet.cell(row=27+e,column=k1+1).value = "No"
            #'else:
            #'    sheet.cell(row=27+e,column=k1+1).value = "Yes"
                                                        
            Ploop+=10 #adds 10 to the pressure until it gets near or close to 100 bar
            e+=1
            
            #below is the code for the graph, the graph code unfortunately doesn't work
            """      
            import xlsxwriter
            pressure_data = [random.random() for _ in range(10)]
            density_data = [random.random() for _ in range(10)]
            # Data location inside excel
            workbook = xlsxwriter.Workbook("Project3_Group8 - Tara McIntosh - Sravya Dhavala.xlsx")
            data_start_loc = [2, 2] # xlsxwriter rquires list, no tuple
            data_end_loc = [data_start_loc[0] + len(random_data), 0]
            chart = workbook.add_chart({'type': 'line'})
            chart.set_y_axis({'name': 'Pressure '})
            chart.set_x_axis({'name': 'Density'})
            chart.set_title({'name': 'Pressure vs Density Chart'})
 
            worksheet = workbook.sheet1()
            # A chart requires data to reference data inside excel
            worksheet.write_column(data_start_loc, data=random_data)
            # The chart needs to explicitly reference data
            chart.add_series({
                    'values': ["Project3_Group8 - Tara McIntosh - Sravya Dhavala.xlsx"] + 'A1' + 'D1',
                    'name': "Pressure and Density Data
                    })
            worksheet.insert_chart('B1', chart)
            workbook.close()  # Write to file """
    
                       
    
    "This section is now just for formatting the columns"  
    #column 1
    
    sheet.cell(row=1,column=1).value = 'Tara McIntosh'
    sheet.cell(row=2,column=1).value = 'Tstart (C)='
    sheet.cell(row=3,column=1).value ='delta T (C) ='
    sheet.cell(row=4,column=1).value = '# of Isotherms ='
    sheet.cell(row=5,column=1).value = 'R ='
    sheet.cell(row=7,column=1).value = 'Name'
    sheet.cell(row=8,column=1).value ='Tc(C) ='
    sheet.cell(row=9,column=1).value = 'Pc(bar)'
    sheet.cell(row=10,column=1).value = 'w ='
    sheet.cell(row=11,column=1).value = 'Mw (g/mol) ='
   # sheet.cell(row=15,column=1).value = 'T'
    
                       #Column 2 info
    sheet.cell(row=1,column=2).value = 'PREOS'
    sheet.cell(row=2,column=2).value= T
    sheet.cell(row=15,column=4).value = 'Pressure (bar)'
    sheet.cell(row=16,column=4).value = 5
    sheet.cell(row=17,column=4).value = 10
    sheet.cell(row=18,column=4).value = 20
    sheet.cell(row=15,column=5).value = 'FugacityV Methane'
    #sheet.cell(row=16,column=5).value = FugacityM
    #sheet.cell(row=17,column=5).value = FVM
    
    sheet.cell(row=15,column=6).value = 'FugacityV Ethane'
    sheet.cell(row=15,column=7).value = 'FugacityV Propane'
    
    sheet.cell(row=3,column=2).value= 1
    sheet.cell(row=4,column=2).value= n
    sheet.cell(row=5,column=2).value= R
    #sheet.cell(row=7,column=2).value= C2
    #sheet.cell(row=8,column=2).value= Tc
    #sheet.cell(row=9,column=2).value= Pc
    #sheet.cell(row=10,column=2).value= w
    #sheet.cell(row=11,column=2).value= Mw
    sheet.cell(row=15,column=2).value = 'Pressure (bar)'
  
    #Column 3 info
    sheet.cell(row=2,column=3).value = 'Pstart (bar)='
    sheet.cell(row=3,column=3).value = 'Pstop='
    sheet.cell(row=4,column=3).value = 'DP (bar)='
    sheet.cell(row=5,column=3).value = '(cm3*bar/(mol*K))'
   # sheet.cell(row=15,column=3).value= C3
    #Column 4 info
    sheet.cell(row=2,column=4).value= P
    sheet.cell(row=3,column=4).value= P+90
    sheet.cell(row=4,column=4).value= 10

    #this next section is for copying the worksheet and placing it directly after the PREOS worksheet
   # from openpyxl.worksheet.copier import WorksheetCopy
   # if C1 in ['Butane','butane']:
   #    C4 = "Butane "
   # else: 
    #  if C1 in ['Oxygen','oxygen']: #inputs compound data
     #     C4 = "Oxygen " 
    #C5 = str(T) + " oC" #adds units to the temperature starts 
   # C52 = str(Tf-1) + " oC"
    #C6 = str(C4)
   # C8 = " to "
    #C7 = C6 + C5 +C8 +C52
   # new_worksheet = project3.create_sheet(str(C7))
   # instance = WorksheetCopy(sheet, new_worksheet)
   # WorksheetCopy.copy_worksheet(instance) 
    
    #now the first worksheet is cleared so that it may be copied over and over for as many times as the program is run 
   #for row in project3['PREOS Data']:
     #   for cell in row:
     #       cell.value = None
    
    
    #these next 2 lines are calling a subroutine to save the worksheet
    k = project3.save("McIntosh.xlsx") #defines k 
    project3 = save(k)   
    
    
if __name__ == '__main__': #allows main routine 'main' to be executed and ended
    main()
